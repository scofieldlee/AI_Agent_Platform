"""
Excel/CSV loader: parse spreadsheet bytes into row-level records for RAG indexing.

Strategy:
    - First non-empty row is treated as the header (column names).
    - Each subsequent non-empty row becomes one retrieval unit (chunk):
      formatted as "列名: 值" lines so embeddings keep the field semantics.

Supported formats: .xlsx / .xlsm (openpyxl), .csv (utf-8-sig).
Legacy .xls is NOT supported (openpyxl limitation) — ask user to convert.
"""

import csv
import hashlib
import io
import logging
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
MAX_FILE_SIZE_MB = 20
MAX_ROWS_PER_SHEET = 20000
MAX_CELL_LENGTH = 2000


def parse_spreadsheet(filename: str, content: bytes) -> Dict:
    """Parse xlsx/xlsm/csv bytes into structured sheet/row records.

    Returns:
        {
            "filename": str,
            "content_hash": md5 hex of raw bytes,
            "sheets": [
                {
                    "name": str,
                    "headers": List[str],
                    "rows": [List[Optional[str]], ...],   # values aligned with headers
                }
            ],
            "total_rows": int,
        }

    Raises:
        ValueError on unsupported extension, decode error, or empty workbook.
    """
    from pathlib import Path

    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"不支持的文件类型 {ext}，仅支持 {', '.join(sorted(ALLOWED_EXTENSIONS))}"
            "（旧版 .xls 请先用 Excel 另存为 .xlsx）"
        )

    content_hash = hashlib.md5(content).hexdigest()

    if ext == ".csv":
        sheets = _parse_csv(content)
    else:
        sheets = _parse_xlsx(content)

    non_empty = [s for s in sheets if s["rows"]]
    if not non_empty:
        raise ValueError("文件中没有可导入的数据行")

    total_rows = sum(len(s["rows"]) for s in non_empty)
    return {
        "filename": filename,
        "content_hash": content_hash,
        "sheets": non_empty,
        "total_rows": total_rows,
    }


def _parse_xlsx(content: bytes) -> List[Dict]:
    """Parse xlsx/xlsm bytes with openpyxl (read-only mode, formulas as cached values)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            raw_rows: List[List[Optional[str]]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_ROWS_PER_SHEET:
                    logger.warning(f"Sheet '{ws.title}' truncated at {MAX_ROWS_PER_SHEET} rows")
                    break
                values = [_clean_cell(v) for v in row]
                if any(v for v in values):  # skip fully empty rows
                    raw_rows.append(values)

            if not raw_rows:
                continue

            headers, data_rows = _extract_headers(raw_rows)
            sheets.append({"name": ws.title or "Sheet", "headers": headers, "rows": data_rows})
        return sheets
    finally:
        wb.close()


def _parse_csv(content: bytes) -> List[Dict]:
    """Parse CSV bytes (utf-8-sig, fallback gbk for Excel-exported Chinese files)."""
    text: str
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("gbk", errors="replace")

    raw_rows: List[List[Optional[str]]] = []
    reader = csv.reader(io.StringIO(text))
    for i, row in enumerate(reader):
        if i >= MAX_ROWS_PER_SHEET:
            logger.warning(f"CSV truncated at {MAX_ROWS_PER_SHEET} rows")
            break
        values = [_clean_cell(v) for v in row]
        if any(v for v in values):
            raw_rows.append(values)

    if not raw_rows:
        return []

    headers, data_rows = _extract_headers(raw_rows)
    return [{"name": "CSV", "headers": headers, "rows": data_rows}]


def _extract_headers(raw_rows: List[List[Optional[str]]]) -> "tuple[List[str], List[List[Optional[str]]]]":
    """Take the first row as headers (generating Col1..N fallbacks), rest as data rows."""
    first = raw_rows[0]
    width = max(len(r) for r in raw_rows)
    headers = []
    for i in range(width):
        name = first[i] if i < len(first) and first[i] else f"列{i + 1}"
        headers.append(str(name)[:100])
    # Deduplicate identical header names (openpyxl common with merged cells)
    seen: Dict[str, int] = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}({seen[h] + 1})")
        else:
            seen[h] = 0
            unique_headers.append(h)

    data_rows = [r for r in raw_rows[1:] if any(v for v in r)]
    return unique_headers, data_rows


def _clean_cell(value) -> Optional[str]:
    """Normalize a cell value to a clean string (None for empty)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float):
        # 3.0 -> "3", keep precision otherwise
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).strip()
    if not text:
        return None
    return text[:MAX_CELL_LENGTH]


def row_to_chunk_text(headers: List[str], row: List[Optional[str]], sheet_name: str) -> str:
    """Format one row as "列名: 值" lines for embedding, keeping field semantics."""
    parts = [f"{headers[i]}: {row[i]}" for i in range(len(headers)) if i < len(row) and row[i]]
    return "\n".join(parts)
